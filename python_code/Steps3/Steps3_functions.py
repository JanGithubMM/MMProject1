from __future__ import division
import pygame, time, os, sys
from gpiozero import MCP3008
import RPi.GPIO as GPIO
import math
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import exifread
import random

kleine_kaartjes = [None,None,None,None,None,None]
grote_kaartjes = [[],[],[],[],[],[],[]]
minimum_links = 1.0
minimum_rechts = 1.0

def init_sensors():
    GPIO.setmode(GPIO.BCM)
    GPIO.setup(4, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
    s_l = MCP3008(1)     #druksensor links
    s_r = MCP3008(3)     #druksensor rechts
    return(s_l, s_r)

def init_pygame_and_screen():  #pygame en screen init
    pygame.init()
    windowInfo = pygame.display.Info()
    screen_w = windowInfo.current_w
    screen_h = windowInfo.current_h
    #screen_w = 800                     #voor testen met andere schermresolutie
    #screen_h = 600                     #voor testen met andere schermresolutie
    print(screen_w, screen_h)
    screen = pygame.display.set_mode((screen_w,screen_h))
    pygame.mouse.set_visible(False)
    return (screen, screen_w, screen_h)

def init_variables(screen_h):
    #my_font = pygame.font.SysFont("Arial", int(screen_h/10))
    my_font = pygame.font.Font("/home/pi/Documents/Fonts/Ruler Stencil Regular.ttf", int(screen_h/10))
    return my_font

def datalogger_init():
    if not os.path.isdir("/home/pi/usbdrv/Beweegdata"):
        os.system("mkdir /home/pi/usbdrv/Beweegdata") 
    if os.path.isfile("/home/pi/usbdrv/Beweegdata/Beweegdata_Steps.xlsx"):
        datalogger = openpyxl.load_workbook("/home/pi/usbdrv/Beweegdata/Beweegdata_Steps.xlsx") #opent het bestaande document op de usb-stick
    else:
        datalogger = openpyxl.load_workbook("/home/pi/Documents/Beweegdata/Beweegdata_Steps_Empty.xlsx") #opent een niet ingevuld exceldocument
    datalogger_sheet = datalogger.active
    return datalogger, datalogger_sheet

def init_steps():
    sensors = init_sensors()
    screen, screen_w, screen_h = init_pygame_and_screen()
    my_font = init_variables(screen_h)
    datalogger, datalogger_sheet = datalogger_init()
    foto_generator = fotos_laden(screen_w, screen_h)
    return sensors, screen, screen_w, screen_h, my_font, datalogger, datalogger_sheet, foto_generator

def fotos_laden(screen_w, screen_h):           #fotos inladen en schalen
    screen_ratio = ((screen_w*(4/5)) / screen_h)
    usbStickHasNoPhotos = True
    for file in os.listdir("/home/pi/usbdrv"):
        if file.endswith(".png")|file.endswith(".PNG")|file.endswith(".jpg")|file.endswith(".JPG")|file.endswith(".jpeg")|file.endswith(".JPEG"):
            usbStickHasNoPhotos = False
            image = pygame.image.load("/home/pi/usbdrv/"+file)
            image = rotate_image("/home/pi/usbdrv/"+file, image)
            #image = scale_binnen_grenzen(image, screen_w*(4/5), screen_h)
            image = scale_binnen_grenzen(image, screen_w*(1304/1920), screen_h*(886/1080), smooth=True)
            yield image
    if (usbStickHasNoPhotos):
            for file in os.listdir("/home/pi/Pictures/Voorbeeldfotos"):
                    if file.endswith(".png")|file.endswith(".PNG")|file.endswith(".jpg")|file.endswith(".JPG")|file.endswith(".jpeg")|file.endswith(".JPEG"):
                            image = pygame.image.load("/home/pi/Pictures/Voorbeeldfotos/"+file)
                            image = rotate_image("/home/pi/Pictures/Voorbeeldfotos/"+file, image)
                            #image = scale_binnen_grenzen(image, screen_w*(4/5), screen_h)
                            image = scale_binnen_grenzen(image, screen_w*(1304/1920), screen_h*(886/1080), smooth=True)
                            yield image

def rotate_image(file_path, image):
    tags = exifread.process_file(open(file_path, 'rb'), stop_tag='Orientation')
    for tag in tags.keys():
        if tag == "Image Orientation":
            if (str(tags[tag]) == "Horizontal (normal)"):
                break   #do nothing
            elif (str(tags[tag]) == 'Rotated 180'):
                image = pygame.transform.rotate(image, 180)
                break
            elif (str(tags[tag]) == 'Rotated 90 CCW'):
                image = pygame.transform.rotate(image, 90)
                break
            elif (str(tags[tag]) == 'Rotated 90 CW'):
                image = pygame.transform.rotate(image, 270)
                break
    return image


def kaartjes_scalen(screen_w, screen_h):
    aantalOefeningen = 7
    for file in os.listdir("/home/pi/Documents/Kaartjes_new"):
            if file.endswith(".png")|file.endswith(".PNG"):
                    raw_kaartje = pygame.image.load("/home/pi/Documents/Kaartjes_new/"+file)
                    for oefeningNummer in range(0, aantalOefeningen):
                            if file.startswith("%d" % oefeningNummer):
                                    groot_plaatje = pygame.transform.smoothscale(raw_kaartje,(int (screen_w/2.33),int(pygame.Surface.get_height(raw_kaartje)/pygame.Surface.get_width(raw_kaartje)*(screen_w/2.33))))
                                    grote_kaartjes[oefeningNummer-1].append(groot_plaatje)
                                    if (len(grote_kaartjes[oefeningNummer-1]) == 1):
                                            klein_plaatje = pygame.transform.smoothscale(raw_kaartje,(int (screen_w/4.64),int(pygame.Surface.get_height(raw_kaartje)/pygame.Surface.get_width(raw_kaartje)*(screen_w/4.64))))
                                            kleine_kaartjes[oefeningNummer-1]=klein_plaatje

def menu_steps(sensors, screen, screen_w, screen_h, my_font, oefening_selected, welcome):
    s_l, s_r = sensors
    sensor_trigger = 0.1
    x = oefening_selected
    oefeningNumber = 0
    
    #Audio Init
    welkom_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/welkom_bij_steps.wav")
    goed_gedaan1 = pygame.mixer.Sound("/home/pi/Documents/Audio/goed_gedaan.wav")
    goed_gedaan2 = pygame.mixer.Sound("/home/pi/Documents/Audio/10_keer_gedaan_goed_werk.wav")
    goed_gedaan3 = pygame.mixer.Sound("/home/pi/Documents/Audio/goed_bezig.wav")
    goed_gedaan_special = pygame.mixer.Sound("/home/pi/Documents/Audio/goed_gedaan_barbershop.wav")

    #achtergrond = pygame.image.load("/home/pi/Documents/Photos/Steps_achtergrond_menu.png")
    achtergrond = pygame.image.load("/home/pi/Documents/Instructie/Steps_achtergrond.png")
    opdracht_text = my_font.render("Kies een oefening", 1, (0,0,0))
    opdracht_text = scale_binnen_grenzen(opdracht_text, screen_w*(2/5), screen_h/10, smooth=True)
    oefening_klaar = pygame.image.load("/home/pi/Documents/Instructie/Oefening_Klaar.png")
    stap_beide = pygame.image.load("/home/pi/Documents/Instructie/Stap_beide.png")
    stap_links = pygame.image.load("/home/pi/Documents/Instructie/Stap_links.png")
    stap_rechts = pygame.image.load("/home/pi/Documents/Instructie/Stap_rechts.png")
    achtergrond = pygame.transform.scale(achtergrond, (screen_w, screen_h),)
    oefening_klaar = pygame.transform.scale(oefening_klaar, (screen_w, screen_h),)
    stap_beide = scale_binnen_grenzen(stap_beide, screen_w/6, screen_h/6, smooth=True)
    stap_links = scale_binnen_grenzen(stap_links, screen_w/10, screen_h/10, smooth=True)
    stap_rechts = scale_binnen_grenzen(stap_rechts, screen_w/10, screen_h/10, smooth=True)

    if (welcome):
        pygame.mixer.Sound.play(welkom_audio)

    ##!let op vanaf nu staat oefeningen 1 dus op positie grote_kaartjes[0]. Hierdoor kan oefening 0 dus niet worden toegevoegd. Oefeningen moeten altijd een nummer groter dan 0 hebben.

    kaartjes_scalen(screen_w, screen_h)
    screen.blit(achtergrond, (0,0))
    screen.blit(opdracht_text,(screen_w/2-opdracht_text.get_width()/2,0))

    huidige_kaartje_x = ((screen_w/2)- pygame.Surface.get_width(grote_kaartjes[0][0])/2)
    huidige_kaartje_y = ((screen_h/2)- pygame.Surface.get_height(grote_kaartjes[0][0])/1.75)
    vorige_kaartje_x = ((huidige_kaartje_x/2)-pygame.Surface.get_width(kleine_kaartjes[1])/2)
    klein_kaartje_y = ((screen_h/2)- pygame.Surface.get_height(kleine_kaartjes[0])/1.75)
    volgende_kaartje_x = (((screen_w-(pygame.Surface.get_width(grote_kaartjes[0][0])+huidige_kaartje_x))/2+(pygame.Surface.get_width(grote_kaartjes[0][0])+huidige_kaartje_x))-pygame.Surface.get_width(kleine_kaartjes[0])/2)

    screen.blit(stap_beide, (screen_w/2 - (stap_beide.get_width()/2), screen_h*(88/100) - (stap_beide.get_height()/2)))


    #dit moet 0 zijn omdat oefeningen bij 0 begint
    animatie = 0 #deze variabele reguleert de animatie
    animatie_counter = 0
    sensor_counter = 0
    links = 0
    rechts = 0
    lock_rechts = 0
    lock_links = 0
    keuze_lock = 0
    klaar_counter = 0
    while True:
        klaar_counter = 0
        events = pygame.event.get()
        for event in events:
            if event.type == pygame.KEYDOWN:
                pygame.quit()
                sys.exit()
        if (s_r_read(s_r.value) < sensor_trigger and s_l_read(s_l.value) > sensor_trigger):
                lock_links = 1
                keuze_lock = 0
        elif (s_r_read(s_r.value) > sensor_trigger and s_l_read(s_l.value) <sensor_trigger):
                lock_rechts = 1
                keuze_lock = 0
        elif (s_r_read(s_r.value) < sensor_trigger and s_l_read(s_l.value) < sensor_trigger):
                keuze_lock = 0

        if(s_l_read(s_l.value) < sensor_trigger and lock_links == 1):
                lock_links = 0
                if x > 0:
                        x = x - 1
                        animatie = 0
        if(s_r_read(s_r.value) < sensor_trigger and lock_rechts == 1):
                lock_rechts = 0
                if x < len(kleine_kaartjes)-1:
                        x = x + 1
                        animatie = 0
        
        if (x-1 >= 0 ):
                screen.blit(kleine_kaartjes[x-1], (vorige_kaartje_x,klein_kaartje_y))
                screen.blit(stap_links, (screen_w*(13/100) - (stap_links.get_width()/2), screen_h*(77/100) - (stap_links.get_height()/2)))
        else:
                screen.blit(achtergrond,(vorige_kaartje_x,klein_kaartje_y),(vorige_kaartje_x,klein_kaartje_y,pygame.Surface.get_width(kleine_kaartjes[0]),pygame.Surface.get_height(kleine_kaartjes[0])))
                screen.blit(achtergrond, (screen_w*(13/100) - (stap_links.get_width()/2), screen_h*(77/100) - (stap_links.get_height()/2)), [screen_w*(13/100) - (stap_links.get_width()/2), screen_h*(77/100) - (stap_links.get_height()/2), stap_links.get_width(),stap_links.get_height()])
        if (x+1 < len(kleine_kaartjes)):
                screen.blit(kleine_kaartjes[x+1], (volgende_kaartje_x,klein_kaartje_y))
                screen.blit(stap_rechts, (screen_w*(87/100) - (stap_rechts.get_width()/2), screen_h*(77/100) - (stap_rechts.get_height()/2)))
        else:
                screen.blit(achtergrond, (volgende_kaartje_x,klein_kaartje_y),(volgende_kaartje_x,klein_kaartje_y,pygame.Surface.get_width(kleine_kaartjes[0]),pygame.Surface.get_height(kleine_kaartjes[0])))
                screen.blit(achtergrond, (screen_w*(87/100) - (stap_rechts.get_width()/2), screen_h*(77/100) - (stap_rechts.get_height()/2)), [screen_w*(87/100) - (stap_rechts.get_width()/2), screen_h*(77/100) - (stap_rechts.get_height()/2), stap_rechts.get_width(),stap_rechts.get_height()])
        screen.blit(grote_kaartjes[x][animatie], (huidige_kaartje_x,huidige_kaartje_y))
        #animatie
        animatie_counter += 1
        if(animatie_counter > 12):
                animatie_counter = 0
                if (animatie == len(grote_kaartjes[x])-1):
                        animatie = 0
                else:
                        animatie += 1
            
        pygame.display.update()
        
        #keuze
        if(s_l_read(s_l.value) > sensor_trigger and s_r_read(s_r.value) > sensor_trigger and keuze_lock == 0):
            keuze_lock = 1
            pygame.mixer.stop()
            return x

def oefening_steps(init_info, oefening_chosen):
    sensors, screen, screen_w, screen_h, my_font, foto_generator = init_info
    s_l, s_r = sensors
    houding1_audio, houding2_audio = init_audio(oefening_chosen)
    achtergrond = set_achtergrond(screen, screen_w, screen_h)
    
    photo = select_volgende_foto(foto_generator, screen_w, screen_h)   #selecteerde de eerste foto, omdat deze functie nog niet eerder aangeroepen is.

    oefening = Oefening(oefening_chosen, screen_w, screen_h)
    uitvoeringen_totaal = oefening.aantal_uitvoeringen
    uitvoeringen_gedaan = 0         #De oefening start altijd met 0 uitvoeringen.
    uitvoeringen_gedaan = verhoog_aantal_uitvoeringen(screen, achtergrond, screen_w, screen_h, uitvoeringen_gedaan, uitvoeringen_totaal, my_font, verhoging = 0)

    fase = 1        #De oefeningen starten allemaal in fase 1
    houding = 1     #De oefeningen starten allemaal in houding 1
    pygame.mixer.Sound.play(houding1_audio)
    random_x = set_random(screen_w)
    random_y = set_random(screen_h)
    teller = 0
    photo_pos_x = 0
    #pygame.draw.rect(screen, (255,255,255),(screen_part_pos_x-50,screen_part_pos_y-50,photo.get_width()+100,photo.get_height()+100))
    set_lijst(screen, photo, screen_w, screen_h, random_x, random_y)
    set_blurred_achtergrond_photo(screen, photo, screen_w, screen_h, random_x, random_y)
    
    while (GPIO.input(4)!=1 and uitvoeringen_gedaan < uitvoeringen_totaal):

        check_keys()
        set_sensor_info(screen, achtergrond, oefening, oefening_chosen, houding, s_l, s_r, screen_h, screen_w)   #niet afhankelijk van de fase

        print(fase)
        time.sleep(0.01)
        if(fase == 1):      #houding 1 aanhouden
            set_houding(screen, screen_w, screen_h, achtergrond, 1, oefening)  #afgebeelde houding aanpassen als nodig

            if(oefening.check_houding(oefening_chosen, houding, s_l, s_r)):      #hieruit komt een True of False
                photo_pos_x = set_photo(screen, photo, screen_w, screen_h, teller, oefening.teller_totaal, random_x, random_y)
                print(photo_pos_x)
                teller += 1                
            if(photo_pos_x >= photo.get_width()):         #houding 1 afgerond
                uitvoeringen_gedaan = verhoog_aantal_uitvoeringen(screen, achtergrond, screen_w, screen_h, uitvoeringen_gedaan, uitvoeringen_totaal, my_font)
                vorige_photo = photo
                vorige_random_x = random_x
                vorige_random_y = random_y
                photo = select_volgende_foto(foto_generator, screen_w, screen_h)    #alvast de volgende foto inladen
                set_houding(screen, screen_w, screen_h, achtergrond, 2, oefening)  #afgebeelde houding aanpassen als nodig
                random_x = set_random(screen_w)
                random_y = set_random(screen_h)
                houding = 2         #wisselen naar houding 2
                teller = 0
                photo_pos_x = 0
                fase = 2
                play_sound(houding2_audio)
        elif(fase == 2):    #wisselen naar houding 2
            #iets dat de gebruiker attendeert om te wisselen van houding
            if(oefening.check_houding(oefening_chosen, houding, s_l, s_r)):      #hieruit komt een True of False
                pygame.mixer.stop()
                #screen.blit(achtergrond, (screen_w/5,0), area=[screen_w/5,0,screen_w*(4/5),screen_h])   #achtergrond leeg maken
                set_blurred_achtergrond_photo(screen, vorige_photo, screen_w, screen_h, vorige_random_x, vorige_random_y)
                set_lijst(screen, photo, screen_w, screen_h, random_x, random_y)
                set_blurred_achtergrond_photo(screen, photo, screen_w, screen_h, random_x, random_y)
                fase = 3
        elif(fase == 3):    #houding 2 aanhouden
            set_houding(screen, screen_w, screen_h, achtergrond, 2, oefening)  #afgebeelde houding aanpassen als nodig
            if(oefening.check_houding(oefening_chosen, houding, s_l, s_r)):      #hieruit komt een True of False
                photo_pos_x = set_photo(screen, photo, screen_w, screen_h, teller, oefening.teller_totaal, random_x, random_y)
                print(photo_pos_x)
                teller += 1
            if(photo_pos_x >= photo.get_width()):         #houding 2 afgerond
                uitvoeringen_gedaan = verhoog_aantal_uitvoeringen(screen, achtergrond, screen_w, screen_h, uitvoeringen_gedaan, uitvoeringen_totaal, my_font)
                vorige_photo = photo
                vorige_random_x = random_x
                vorige_random_y = random_y
                photo = select_volgende_foto(foto_generator, screen_w, screen_h)    #alvast de volgende foto inladen
                set_houding(screen, screen_w, screen_h, achtergrond, 1, oefening)  #afgebeelde houding aanpassen als nodig
                random_x = set_random(screen_w)
                random_y = set_random(screen_h)
                houding = 1         #wisselen naar houding 1
                teller = 0
                photo_pos_x = 0
                fase = 4
                play_sound(houding1_audio)
        elif(fase == 4):    #wisselen naar houding 1
            #iets dat de gebruiker attendeert om te wisselen van houding
            if(oefening.check_houding(oefening_chosen, houding, s_l, s_r)):      #hieruit komt een True of False
                pygame.mixer.stop()
                #screen.blit(achtergrond, (screen_w/5,0), area=[screen_w/5,0,screen_w*(4/5),screen_h])   #achtergrond leeg maken
                set_blurred_achtergrond_photo(screen, vorige_photo, screen_w, screen_h, vorige_random_x, vorige_random_y)
                set_lijst(screen, photo, screen_w, screen_h, random_x, random_y)
                set_blurred_achtergrond_photo(screen, photo, screen_w, screen_h, random_x, random_y)
                fase = 1
    oefening_result = (uitvoeringen_gedaan, uitvoeringen_gedaan >= uitvoeringen_totaal)
    pygame.mixer.stop()
    oefening_klaar = pygame.image.load("/home/pi/Documents/Instructie/Oefening_Klaar.png")
    oefening_klaar = pygame.transform.scale(oefening_klaar, (screen_w, screen_h))
    screen.blit(oefening_klaar,(0,0))
    pygame.display.update()
    klaar_counter = 0
    while(klaar_counter < 30):
        time.sleep(0.1)
        if(s_l_read(s_l.value) < 0.1 and s_r_read(s_r.value) <  0.1):
            klaar_counter += 1
        else:
            klaar_counter = 0
    return oefening_result

def set_photo(screen, photo, screen_w, screen_h, teller, teller_totaal, random_x, random_y):
    photo_part_h = photo.get_height()
    photo_part_w = math.ceil(photo.get_width()/teller_totaal)
    photo_part_pos_x = photo_part_w * teller
    photo_part_pos_y = 0
    screen_part_pos_x = screen_w*(1156/1920) - photo.get_width()/2 + (photo_part_w * teller) + random_x
    screen_part_pos_y = screen_h*(537/1080) - photo_part_h/2 + random_y
    
    screen.blit(photo, (screen_part_pos_x,screen_part_pos_y), area=[photo_part_pos_x,photo_part_pos_y,photo_part_w,photo_part_h]) # van links naar rechts
    pygame.display.update()
    return photo_part_pos_x

def set_photo_blur(screen, photo, screen_w, screen_h, teller, teller_totaal):
    photo_original_h = photo.get_height()
    photo_original_w = photo.get_width()
    screen_part_pos_x = screen_w*(1156/1920) - photo.get_width()/2
    screen_part_pos_y = screen_h*(537/1080) - photo.get_height()/2
    
    if(teller >= teller_totaal-1):
        screen.blit(photo, (screen_part_pos_x,screen_part_pos_y))
    else:
        scale = (teller*teller)/(teller_totaal*teller_totaal*3)
        photo_scale_h = int(math.floor(photo_original_h*scale))+1
        photo_scale_w = int(math.floor(photo_original_w*scale))+1
        photo = pygame.transform.scale(photo, (photo_scale_w,photo_scale_h))
        photo = pygame.transform.smoothscale(photo, (photo_original_w,photo_original_h))
        screen.blit(photo, (screen_part_pos_x,screen_part_pos_y))
    
    pygame.display.update()
    return teller 

def set_blurred_achtergrond_photo(screen, photo, screen_w, screen_h, random_x, random_y):
    photo_original_h = photo.get_height()
    photo_original_w = photo.get_width()
    screen_part_pos_x = screen_w*(1156/1920) - photo.get_width()/2 + random_x
    screen_part_pos_y = screen_h*(537/1080) - photo.get_height()/2 + random_y
    photo_scale_h = int(math.floor(photo_original_h/40))+1 # +1 om errors te voorkomen
    photo_scale_w = int(math.floor(photo_original_w/40))+1
    photo = pygame.transform.scale(photo, (photo_scale_w,photo_scale_h))
    photo = pygame.transform.smoothscale(photo, (photo_original_w,photo_original_h))
    screen.blit(photo, (screen_part_pos_x,screen_part_pos_y))

def set_lijst(screen, photo, screen_w, screen_h, random_x, random_y):
    screen_pos_x = screen_w*(1156/1920) - photo.get_width()/2 + random_x
    screen_pos_y = screen_h*(537/1080) - photo.get_height()/2 + random_y

    schaduw = pygame.Surface((photo.get_width()+45,photo.get_height()+45))
    schaduw.set_alpha(100)
    schaduw.fill((130,130,130))

##    lijst = pygame.Surface((photo.get_width()+40,photo.get_height()+40))
##    lijst.set_alpha(255)
##    lijst.fill((255,255,255))
##    schaduw.blit(lijst, (0,0))
##    schaduw = pygame.transform.scale(schaduw, (100,100))
##    schaduw = pygame.transform.smoothscale(schaduw, (photo.get_width()+45,photo.get_height()+45))
    
    screen.blit(schaduw, (screen_pos_x -20,screen_pos_y -20))
    pygame.draw.rect(screen, (255,255,255),(screen_pos_x - 20,screen_pos_y -20,photo.get_width()+40,photo.get_height()+40))
    pygame.draw.rect(screen, (129,130,135),(screen_pos_x ,screen_pos_y,photo.get_width(),photo.get_height()))

def select_foto(volgende_foto):
    image = volgende_foto
    photo = image
    return photo

def set_achtergrond(screen, screen_w, screen_h):
    #achtergrond = pygame.image.load("/home/pi/Documents/Instructie/Steps_oefening_achtergrond.jpg")
    achtergrond = pygame.image.load("/home/pi/Documents/Instructie/Steps_achtergrond_menu.png")
    #achtergrond = pygame.image.load("/home/pi/Documents/Photos/Steps_achtergrond_def.png")
    achtergrond = pygame.transform.scale(achtergrond, (screen_w, screen_h))
    screen.blit(achtergrond, (0,0))
    pygame.display.update()
    return achtergrond


def select_volgende_foto(foto_generator, screen_w, screen_h):
    try:
            volgende_foto = foto_generator.next()
    except StopIteration:
            foto_generator = fotos_laden(screen_w, screen_h)
            volgende_foto = foto_generator.next()

    #photo_pos_x = math.ceil((screen_w*(3/5))-(volgende_foto.get_width()/2))
    return volgende_foto

def verhoog_aantal_uitvoeringen(screen, achtergrond, screen_w, screen_h, aantal_uitvoeringen, uitvoeringen_totaal, my_font, verhoging = 1):
    aantal_uitvoeringen += verhoging
    uitvoering_text = my_font.render("%s van de %s" % (aantal_uitvoeringen, uitvoeringen_totaal), 1, (0,0,0))
    screen.blit(achtergrond, (0,screen_h*(4/5)), area=[0,screen_h*(4/5),screen_w*(1/5),screen_h*(1/5)])
    uitvoering_text = scale_binnen_grenzen(uitvoering_text, screen_w*(1/6), screen_h*(1/6))
    screen.blit(uitvoering_text, (screen_w*(1/10)-uitvoering_text.get_width()/2, screen_h*(9/10)-uitvoering_text.get_height()/2)) #de uitvoeringstekst
    pygame.display.update([0,screen_h*(4/5),screen_w*(1/5),screen_h*(1/5)])
    return aantal_uitvoeringen

class Oefening:
    def __init__(self, oefening_chosen, screen_w, screen_h):
        if(oefening_chosen == 0):
            self.teller_totaal = 150
            self.aantal_uitvoeringen = 10
            self.image_houding1 = pygame.image.load("/home/pi/Documents/Oefeningen/15c_Extensie_knie_in_zit.png")        #aanpasbare afbeelding
            self.image_houding1 = scale_binnen_grenzen(self.image_houding1, screen_w/5, screen_h*(2/5), smooth=True)
            #self.image_houding1 = pygame.transform.smoothscale(self.image_houding1, (int(screen_w/100)*14, int(screen_w/100)*20))
            self.image_houding2 = pygame.image.load("/home/pi/Documents/Oefeningen/15b_Extensie_knie_in_zit.png")        #aanpasbare afbeelding
            self.image_houding2 = scale_binnen_grenzen(self.image_houding2, screen_w/5, screen_h*(2/5),smooth=True)
            #self.image_houding2 = pygame.transform.smoothscale(self.image_houding2, (int(screen_w/100)*14, int(screen_w/100)*20))
        elif(oefening_chosen == 1):
            self.teller_totaal = 50
            self.aantal_uitvoeringen = 10
            self.image_houding1 = pygame.image.load("/home/pi/Documents/Oefeningen/14a_Zit_naar_stand_twee_handen.png")        #aanpasbare afbeelding
            #self.image_houding1 = scale_binnen_grenzen(self.image_houding1, screen_w/5, screen_h*(2/5), smooth=True)
            self.image_houding1 = pygame.transform.smoothscale(self.image_houding1, (int(screen_w/100)*14, int(screen_w/100)*20))
            self.image_houding2 = pygame.image.load("/home/pi/Documents/Oefeningen/14b_Zit_naar_stand_twee_handen.png")        #aanpasbare afbeelding
            #self.image_houding2 = scale_binnen_grenzen(self.image_houding2, screen_w/5, screen_h*(2/5),smooth=True)
            self.image_houding2 = pygame.transform.smoothscale(self.image_houding2, (int(screen_w/100)*14, int(screen_w/100)*20))
        elif(oefening_chosen == 2):
            self.teller_totaal = 150
            self.aantal_uitvoeringen = 10
            self.image_houding1 = pygame.image.load("/home/pi/Documents/Oefeningen/8a_achteruit_lopen_zonder.png")        #aanpasbare afbeelding
            #self.image_houding1 = scale_binnen_grenzen(self.image_houding1, screen_w/5, screen_h*(2/5), smooth=True)
            self.image_houding1 = pygame.transform.smoothscale(self.image_houding1, (int(screen_w/100)*14, int(screen_w/100)*20))
            self.image_houding2 = pygame.image.load("/home/pi/Documents/Oefeningen/8b_achteruit_lopen_zonder.png")        #aanpasbare afbeelding
            #self.image_houding2 = scale_binnen_grenzen(self.image_houding2, screen_w/5, screen_h*(2/5),smooth=True)
            self.image_houding2 = pygame.transform.smoothscale(self.image_houding2, (int(screen_w/100)*14, int(screen_w/100)*20))
        elif(oefening_chosen == 3):
            self.teller_totaal = 50
            self.aantal_uitvoeringen = 10
            self.image_houding1 = pygame.image.load("/home/pi/Documents/Oefeningen/2a_Been_zijwaarts_heffen.png")        #aanpasbare afbeelding
            #self.image_houding1 = scale_binnen_grenzen(self.image_houding1, screen_w/5, screen_h*(2/5), smooth=True)
            self.image_houding1 = pygame.transform.smoothscale(self.image_houding1, (int(screen_w/100)*14, int(screen_w/100)*20))
            self.image_houding2 = pygame.image.load("/home/pi/Documents/Oefeningen/2c_Been_zijwaarts_heffen.png")        #aanpasbare afbeelding
            #self.image_houding2 = scale_binnen_grenzen(self.image_houding2, screen_w/5, screen_h*(2/5),smooth=True)
            self.image_houding2 = pygame.transform.smoothscale(self.image_houding2, (int(screen_w/100)*14, int(screen_w/100)*20))
        elif(oefening_chosen == 4):
            self.teller_totaal = 150
            self.aantal_uitvoeringen = 10
            self.image_houding1 = pygame.image.load("/home/pi/Documents/Oefeningen/1a_Hiel_naar_de_bil.png")        #aanpasbare afbeelding
            #self.image_houding1 = scale_binnen_grenzen(self.image_houding1, screen_w/5, screen_h*(2/5), smooth=True)
            self.image_houding1 = pygame.transform.smoothscale(self.image_houding1, (int(screen_w/100)*14, int(screen_w/100)*20))
            self.image_houding2 = pygame.image.load("/home/pi/Documents/Oefeningen/1e_Hiel_naar_de_bil.png")        #aanpasbare afbeelding
            #self.image_houding2 = scale_binnen_grenzen(self.image_houding2, screen_w/5, screen_h*(2/5),smooth=True)
            self.image_houding2 = pygame.transform.smoothscale(self.image_houding2, (int(screen_w/100)*14, int(screen_w/100)*20))
        elif(oefening_chosen == 5):
            self.teller_totaal = 150
            self.aantal_uitvoeringen = 10
            self.image_houding1 = pygame.image.load("/home/pi/Documents/Oefeningen/12c_Tien_seconden_op_1_been_staan_met_steun.png")        #aanpasbare afbeelding
            #self.image_houding1 = scale_binnen_grenzen(self.image_houding1, screen_w/5, screen_h*(2/5), smooth=True)
            self.image_houding1 = pygame.transform.smoothscale(self.image_houding1, (int(screen_w/100)*14, int(screen_w/100)*20))
            self.image_houding2 = pygame.image.load("/home/pi/Documents/Oefeningen/12a_Tien_seconden_op_1_been_staan_met_steun.png")        #aanpasbare afbeelding
            #self.image_houding2 = scale_binnen_grenzen(self.image_houding2, screen_w/5, screen_h*(2/5),smooth=True)
            self.image_houding2 = pygame.transform.smoothscale(self.image_houding2, (int(screen_w/100)*14, int(screen_w/100)*20))
        elif(oefening_chosen == 6):
            self.teller_totaal = 150
            self.aantal_uitvoeringen = 10
            self.image_houding1 = pygame.image.load("/home/pi/Documents/Oefeningen/13c_Tien_seconden_op_1_been_staan_zonder_steun.png")        #aanpasbare afbeelding
            #self.image_houding1 = scale_binnen_grenzen(self.image_houding1, screen_w/5, screen_h*(2/5), smooth=True)
            self.image_houding1 = pygame.transform.smoothscale(self.image_houding1, (int(screen_w/100)*14, int(screen_w/100)*20))
            self.image_houding2 = pygame.image.load("/home/pi/Documents/Oefeningen/13a_Tien_seconden_op_1_been_staan_zonder_steun.png")        #aanpasbare afbeelding
            #self.image_houding2 = scale_binnen_grenzen(self.image_houding2, screen_w/5, screen_h*(2/5),smooth=True)
            self.image_houding2 = pygame.transform.smoothscale(self.image_houding2, (int(screen_w/100)*14, int(screen_w/100)*20))
    def check_houding(self, oefening_chosen, houding, s_l, s_r):
        if(houding == 1):
            s_trigger_links = 0.1
            s_trigger_rechts = 0.1
            if(oefening_chosen == 0):       #knie extensie
                s_trigger_links = 0.1
                s_trigger_rechts = 0.1
                return(s_l_read(s_l.value) < s_trigger_links and s_r_read(s_r.value) > s_trigger_rechts)
            elif(oefening_chosen == 1):     #staanzitten
                s_trigger_total = 0.8
                return(s_l_read(s_l.value) + s_r_read(s_r.value) < s_trigger_total)
            elif(oefening_chosen == 2):     #achterenlopen
                s_trigger_total = s_l_read(s_l.value) + s_r_read(s_r.value)
                return(s_l_read(s_l.value) > (s_trigger_total*0.7) and s_trigger_total > 0.1)
            elif(oefening_chosen == 3):     #beenheffen
                return(s_l_read(s_l.value) > s_trigger_links and s_r_read(s_r.value) < s_trigger_rechts)
            elif(oefening_chosen == 4):     #hak naar bil
                return(s_l_read(s_l.value) > s_trigger_links and s_r_read(s_r.value) < s_trigger_rechts)
            elif(oefening_chosen == 5 or oefening_chosen == 6):
                return(s_l_read(s_l.value) < s_trigger_links and s_r_read(s_r.value) > s_trigger_rechts)
        elif(houding == 2):
            s_trigger_links = 0.1
            s_trigger_rechts = 0.1
            if(oefening_chosen == 0):       #knie extensie
                s_trigger_links = 0.1
                s_trigger_rechts = 0.1
                return(s_l_read(s_l.value) > s_trigger_links and s_r_read(s_r.value) < s_trigger_rechts)
            elif(oefening_chosen == 1):     #staanzitten
                s_trigger_total = 0.8
                return(s_l_read(s_l.value) + s_r_read(s_r.value) > s_trigger_total)
            elif(oefening_chosen == 2):     #achterenlopen
                s_trigger_total = s_l_read(s_l.value) + s_r_read(s_r.value)
                return(s_r_read(s_r.value) > (s_trigger_total*0.7) and s_trigger_total > 0.1)
            elif(oefening_chosen == 3):     #beenheffen
                return(s_l_read(s_l.value) < s_trigger_links and s_r_read(s_r.value) > s_trigger_rechts)
            elif(oefening_chosen == 4):     #hak naar bil
                return(s_l_read(s_l.value) < s_trigger_links and s_r_read(s_r.value) > s_trigger_rechts)
            elif(oefening_chosen == 5 or oefening_chosen == 6):
                return(s_l_read(s_l.value) > s_trigger_links and s_r_read(s_r.value) < s_trigger_rechts)

def set_houding(screen, screen_w, screen_h, achtergrond, houding, oefening):    
    if(houding == 1):   #de houding die afgebeeld wordt
        uitleg_plaatje = oefening.image_houding1
    elif(houding == 2):  #de houding die afgebeeld wordt
        uitleg_plaatje = oefening.image_houding2
    houding_x = screen_w*(1/10)-uitleg_plaatje.get_width()/2
    houding_y = screen_h*(1/5)-uitleg_plaatje.get_height()/2
    screen.blit(achtergrond, (houding_x, houding_y), area=[houding_x,houding_y,uitleg_plaatje.get_width(),uitleg_plaatje.get_height()])
    screen.blit(uitleg_plaatje, (houding_x,houding_y))
    pygame.display.update()

def set_sensor_info(screen, achtergrond, oefening, oefening_chosen, houding, s_l, s_r, screen_h, screen_w):
        
    ### schalend maken op het scherm!!  ###
    #breedte = screen_w/5
    #hoogte = screen_h*(2/5)
    #x = 0
    #y = screen_h*(2/5)
    
    #hoogte_sensor_links = math.floor(s_l_read(s_l.value) * 100)
    #hoogte_sensor_rechts = math.floor(s_r_read(s_r.value) * 100)

    hoogte_sensor_links = s_l_read(s_l.value) * (screen_h*(2/5))
    hoogte_sensor_rechts = s_r_read(s_r.value) * (screen_h*(2/5))

    if(oefening.check_houding(oefening_chosen, houding, s_l, s_r)):
        kleur_sensor = (50,255,50)
    else:
        kleur_sensor = (255,50,50)
    
    screen.blit(achtergrond, (0,screen_h*(2/5)), area=[0,screen_h*(2/5),screen_w*(1/5),screen_h*(2/5)])
    #voetstap_links = pygame.image.load("/home/pi/Downloads/Voetstap_links.png")
    #voetstap_links = scale_binnen_grenzen(voetstap_links, screen_w*(1/10), screen_h*(2/5), smooth=False)
    #screen.blit(voetstap_links, (0,screen_h*(2/5)))
    
    if(hoogte_sensor_links > 10):
        pygame.draw.rect(screen,(kleur_sensor),(screen_w*(1/50),math.floor(screen_h*(4/5))-hoogte_sensor_links,screen_w*(3/50),hoogte_sensor_links))
    if(hoogte_sensor_rechts > 10):
        pygame.draw.rect(screen,(kleur_sensor),(screen_w*(6/50),math.floor(screen_h*(4/5))-hoogte_sensor_rechts,screen_w*(3/50),hoogte_sensor_rechts))
    pygame.display.update((0,screen_h*(2/5),screen_w*(1/5),screen_h*(2/5)))

    ######
    #pygame.display.update((0,screen_h*(2/5),screen_w/5,screen_h*(2/5)))
    ######
    

def init_audio(oefening_chosen):
    if(oefening_chosen == 0):
        houding1_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/strek_linkerbeen.wav")
        houding2_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/strek_rechterbeen.wav")
    elif(oefening_chosen == 1):
        houding1_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/ga_zitten.wav")
        houding2_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/ga_staan.wav")
    elif(oefening_chosen == 2):
        houding1_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/leun_naar_voren.wav")
        houding2_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/leun_naar_achter.wav")
    elif(oefening_chosen == 3):
        houding1_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/til_rechterbeen_op.wav")
        houding2_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/til_linkerbeen_op.wav")
    elif(oefening_chosen == 4):
        houding1_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/buig_rechterknie.wav")
        houding2_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/buig_linkerknie.wav")
    elif(oefening_chosen == 5 or oefening_chosen == 6):
        houding1_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/sta_op_rechterbeen.wav")
        houding2_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/sta_op_linkerbeen.wav")
    return houding1_audio, houding2_audio

def play_sound(houding_audio):
    #time.sleep(1)
    pygame.mixer.Sound.play(houding_audio)

def check_keys():
    events = pygame.event.get()
    for event in events:
        if event.type == pygame.KEYDOWN:
            pygame.quit()
            sys.exit(0)

def s_r_read(value):
    global minimum_rechts
    minimum_rechts = minimum_rechts + (minimum_rechts/2500)     #om te voorkomen dat er 1 keer een lage waarde wordt gemeten.
    if value < minimum_rechts:
        minimum_rechts = value
    return (value - minimum_rechts)

def s_l_read(value):
    global minimum_links
    minimum_links = minimum_links + (minimum_links/2500)        #om te voorkomen dat er 1 keer een lage waarde wordt gemeten.
    if value < minimum_links:
        minimum_links = value
    return (value - minimum_links)

def scale_binnen_grenzen(image, vak_w, vak_h, smooth=False):
    vak_ratio = vak_w/vak_h
    image_ratio = image.get_width()/image.get_height()
    
    if image_ratio < vak_ratio:          #balk aan de zijkant
        if(smooth):
            image = pygame.transform.smoothscale(image, (int(vak_h*image_ratio), int(vak_h)))
        else:
            image = pygame.transform.scale(image, (int(vak_h*image_ratio), int(vak_h)))
    else:                                #balk aan de onder en bovenkant
        if(smooth):
            image = pygame.transform.smoothscale(image, (int(vak_w), int(vak_w/image_ratio)))
        else:
            image = pygame.transform.scale(image, (int(vak_w), int(vak_w/image_ratio)))
    return image

def set_random(screen_w_or_h):
    random_x_or_y = random.randrange(-int(screen_w_or_h/40), int(screen_w_or_h/40))
    return random_x_or_y


    ###     excel       ###

def excel_start_oefening():
    return datetime.now()

def excel_save(datalogger, datalogger_sheet, start_tijd_oefening, oefening_result, oefening_chosen):
    (uitvoeringen_gedaan, gehaald) = oefening_result

    #(oefening = "Extensie knie", tijdOefening = "1 min", aantalKeer = 0, afgerond = "ja", moeilijkheid = 1, client = "client"):

    rowCounter = 2
    while datalogger_sheet.cell(row=rowCounter, column=1).value != None:
            rowCounter += 1
            
    datalogger_sheet.cell(row=rowCounter, column=1).value = datetime.now().strftime('%d-%m-%Y')
    datalogger_sheet.cell(row=rowCounter, column=2).value = datetime.now().strftime('%H:%M')
    datalogger_sheet.cell(row=rowCounter, column=3).value = get_oefening_naam(oefening_chosen)
    datalogger_sheet.cell(row=rowCounter, column=4).value = (datetime.now() - start_tijd_oefening)
    datalogger_sheet.cell(row=rowCounter, column=5).value = uitvoeringen_gedaan
    if gehaald:
            datalogger_sheet.cell(row=rowCounter, column=6).fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type = 'solid')
    else:
            datalogger_sheet.cell(row=rowCounter, column=6).fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = "solid")
    #datalogger_sheet.cell(row=rowCounter, column=7).value = moeilijkheid
    #datalogger_sheet.cell(row=rowCounter, column=8).value = client
    datalogger.save("/home/pi/usbdrv/Beweegdata/Beweegdata_Steps.xlsx")
    datalogger.save("/home/pi/Documents/Beweegdata/Beweegdata_Steps.xlsx") #Slaat op als root in niet-startx modus en overwrite files van vorige usb-sticks

def convert_timedelta(duration):
	seconds = duration.seconds
	minutes = (seconds % 3600) // 60
	seconds = (seconds % 60)
	return "{} min en {} sec".format(minutes, seconds)



def get_oefening_naam(oefening_nummer):
    if(oefening_nummer == 0):
        oefening_naam = "Knie extensie"
    elif(oefening_nummer == 1):
        oefening_naam = "Staan zitten"
    elif(oefening_nummer == 2):
        oefening_naam = "Achteren lopen"
    elif(oefening_nummer == 3):
        oefening_naam = "Been heffen"
    elif(oefening_nummer == 4):
        oefening_naam = "Hak naar bil"
    elif(oefening_nummer == 5 or oefening_nummer == 6):
        oefening_naam = "Op 1 been staan"
    return oefening_naam


def oefening_uitleg(sensors, screen, screen_w, screen_h, my_font, oefening_chosen):

    if(oefening_chosen == 0):
        #image
        instructie = pygame.image.load("/home/pi/Documents/Instructie/Instructie_ExtensieKnie.png")
        instructie = pygame.transform.scale(instructie, (screen_w, screen_h))
        plaatsing_achtergrond = pygame.image.load("/home/pi/Documents/Instructie/Instructie_StepsHorizontaal.png")
        plaatsing_achtergrond = pygame.transform.scale(plaatsing_achtergrond, (screen_w, screen_h))
        mannetje = pygame.image.load("/home/pi/Documents/Instructie/Oudere_zit.png")
        mannetje_ratio = mannetje.get_width()/mannetje.get_height()
        mannetje = pygame.transform.smoothscale(mannetje,(int(screen_w/5), int(screen_w/5/mannetje_ratio)))
        pos_mannetje = ((screen_w/2 - screen_w/9),(screen_h/2))
        #audio
        uitleg_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/uitleg_benen_strekken.wav")
        plaatsing_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/plaatsing_benen_strekken.wav")
    elif(oefening_chosen == 1):
        #image
        instructie = pygame.image.load("/home/pi/Documents/Instructie/Instructie_StaanZitten.png")
        instructie = pygame.transform.scale(instructie, (screen_w, screen_h))
        plaatsing_achtergrond = pygame.image.load("/home/pi/Documents/Instructie/Instructie_StepsHorizontaal.png")
        plaatsing_achtergrond = pygame.transform.scale(plaatsing_achtergrond, (screen_w, screen_h))
        mannetje = pygame.image.load("/home/pi/Documents/Instructie/Oudere_zit.png")
        mannetje_ratio = mannetje.get_width()/mannetje.get_height()
        mannetje = pygame.transform.smoothscale(mannetje,(int(screen_w/5), int(screen_w/5/mannetje_ratio)))
        pos_mannetje = ((screen_w/2 - screen_w/9),(screen_h/2))
        #audio
        uitleg_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/uitleg_staan_zitten.wav")
        plaatsing_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/plaatsing_staan_zitten.wav")
    elif(oefening_chosen == 2):
        #image
        instructie = pygame.image.load("/home/pi/Documents/Instructie/Instructie_AchterenLopen.png")
        instructie = pygame.transform.scale(instructie, (screen_w, screen_h))
        plaatsing_achtergrond = pygame.image.load("/home/pi/Documents/Instructie/Instructie_StepsVerticaal.png")
        plaatsing_achtergrond = pygame.transform.scale(plaatsing_achtergrond, (screen_w, screen_h))
        mannetje = pygame.image.load("/home/pi/Documents/Instructie/Oudere_staat1.png")
        mannetje_ratio = mannetje.get_width()/mannetje.get_height()
        mannetje = pygame.transform.smoothscale(mannetje,(int(screen_w/3), int(screen_w/5/mannetje_ratio)))
        pos_mannetje = ((screen_w/2 - screen_w/5.5),(screen_h/2-screen_h/6))
        #audio
        uitleg_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/uitleg_achteren_lopen.wav")
        plaatsing_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/plaatsing_achteren_lopen.wav")
    elif(oefening_chosen == 3):
        #image
        instructie = pygame.image.load("/home/pi/Documents/Instructie/Instructie_BeenHeffen.png")
        instructie = pygame.transform.scale(instructie, (screen_w, screen_h))
        plaatsing_achtergrond = pygame.image.load("/home/pi/Documents/Instructie/Instructie_StepsHorizontaal.png")
        plaatsing_achtergrond = pygame.transform.scale(plaatsing_achtergrond, (screen_w, screen_h))
        mannetje = pygame.image.load("/home/pi/Documents/Instructie/Oudere_staat2.png")
        mannetje_ratio = mannetje.get_width()/mannetje.get_height()
        mannetje = pygame.transform.smoothscale(mannetje,(int(screen_w/3), int(screen_w/5/mannetje_ratio)))
        pos_mannetje = ((screen_w/2 - screen_w/5.5),(screen_h/2-screen_h/6))
        #audio
        uitleg_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/uitleg_been_heffen.wav")
        plaatsing_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/plaatsing_been_heffen.wav")            
    elif(oefening_chosen == 4):
        #image
        instructie = pygame.image.load("/home/pi/Documents/Instructie/Instructie_HakNaarBil.png")
        instructie = pygame.transform.scale(instructie, (screen_w, screen_h))
        plaatsing_achtergrond = pygame.image.load("/home/pi/Documents/Instructie/Instructie_StepsHorizontaal.png")
        plaatsing_achtergrond = pygame.transform.scale(plaatsing_achtergrond, (screen_w, screen_h))
        mannetje = pygame.image.load("/home/pi/Documents/Instructie/Oudere_staat2.png")
        mannetje_ratio = mannetje.get_width()/mannetje.get_height()
        mannetje = pygame.transform.smoothscale(mannetje,(int(screen_w/3), int(screen_w/5/mannetje_ratio)))
        pos_mannetje = ((screen_w/2 - screen_w/5.5),(screen_h/2-screen_h/6))
        #audio
        uitleg_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/uitleg_hak_naar_bil.wav")
        plaatsing_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/plaatsing_hak_naar_bil.wav")
    elif(oefening_chosen == 5 or oefening_chosen == 6):
        #image
        instructie = pygame.image.load("/home/pi/Documents/Instructie/Instructie_OpEenBeenStaan.png")
        instructie = pygame.transform.scale(instructie, (screen_w, screen_h))
        plaatsing_achtergrond = pygame.image.load("/home/pi/Documents/Instructie/Instructie_StepsHorizontaal.png")
        plaatsing_achtergrond = pygame.transform.scale(plaatsing_achtergrond, (screen_w, screen_h))
        mannetje = pygame.image.load("/home/pi/Documents/Instructie/Oudere_staat2.png")
        mannetje_ratio = mannetje.get_width()/mannetje.get_height()
        mannetje = pygame.transform.smoothscale(mannetje,(int(screen_w/3), int(screen_w/5/mannetje_ratio)))
        pos_mannetje = ((screen_w/2 - screen_w/5.5),(screen_h/2-screen_h/6))
        #audio
        uitleg_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/uitleg_op_een_been_staan.wav")
        plaatsing_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/plaatsing_op_een_been_staan.wav")

    #video
    pygame.mixer.Sound.play(uitleg_audio)

    screen.blit(instructie, (0,0))
    pygame.display.update()    
    for i in range (0, 100):
        time.sleep(uitleg_audio.get_length()/100)
        if (GPIO.input(4) == 1):
            pygame.mixer.stop()
            return

    pygame.mixer.Sound.play(plaatsing_audio)

    screen.blit(plaatsing_achtergrond, (0,0))
    pygame.display.update()
    for i in range (0, 100):
        time.sleep(plaatsing_audio.get_length()/2 /100)
        if (GPIO.input(4) == 1):
            pygame.mixer.stop()
            return
    screen.blit(mannetje, pos_mannetje)
    pygame.display.update()
    for i in range (0, 100):
        time.sleep(plaatsing_audio.get_length()/2 /100)
        if (GPIO.input(4) == 1):
            pygame.mixer.stop()
            
